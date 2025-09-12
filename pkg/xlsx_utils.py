# -*- coding: utf-8 -*-
"""Utility helpers for XLSX report generation."""
from __future__ import annotations

from typing import Iterable, Dict, List

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Common fills used across reports
GREEN = PatternFill(start_color="E7F7E7", end_color="E7F7E7", fill_type="solid")
RED = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def _autosize(ws) -> None:
    """Automatically set column widths based on cell contents."""
    max_width: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for idx, val in enumerate(row, start=1):
            s = "" if val is None else str(val)
            width = max(3, min(120, int(len(s) * 1.1) + 2))
            if idx not in max_width or width > max_width[idx]:
                max_width[idx] = width
    for idx, w in max_width.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

def _borders(ws) -> None:
    """Apply thin gray borders to all cells of the sheet."""
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

def _style(ws, left_cols: Iterable[int], yes_no_cols: Iterable[str], status_col: str) -> None:
    """Apply common styling and conditional formatting to report sheets.

    Args:
        ws: Worksheet to style.
        left_cols: Columns (1-indexed) that should use left alignment.
        yes_no_cols: Columns (letters) where "Да"/"Нет" conditional formatting applies.
        status_col: Column letter containing the overall status field.
    """
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = GRAY

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in left_cols:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col in yes_no_cols:
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}", CellIsRule(operator='equal', formula=['"Да"'], fill=GREEN)
        )
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}", CellIsRule(operator='equal', formula=['"Нет"'], fill=RED)
        )

    ws.conditional_formatting.add(
        f"{status_col}2:{status_col}{ws.max_row}", CellIsRule(operator='equal', formula=['"OK"'], fill=GREEN)
    )
    ws.conditional_formatting.add(
        f"{status_col}2:{status_col}{ws.max_row}", CellIsRule(operator='notEqual', formula=['"OK"'], fill=RED)
    )

def create_summary_sheet(wb, rows: List[Dict], status_key: str = "Статус") -> Dict[str, int]:
    """Create a standard Summary sheet and return statistics.

    Args:
        wb: Workbook where the sheet will be created.
        rows: Data rows containing status information.
        status_key: Key in row dicts holding overall status.

    Returns:
        Dictionary with total, ok and errors counts.
    """
    sm = wb.create_sheet("Summary")
    total = len(rows)
    ok = sum(1 for r in rows if r.get(status_key) == "OK")
    errors = total - ok
    sm.append(["Метрика", "Значение"])
    sm.append(["Всего строк", total])
    sm.append(["OK", ok])
    sm.append(["Ошибки (все не-OK)", errors])
    sm.auto_filter.ref = sm.dimensions
    sm.freeze_panes = "A2"
    for c in sm[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = GRAY
    _autosize(sm)
    _borders(sm)
    return {"total": total, "ok": ok, "errors": errors}
