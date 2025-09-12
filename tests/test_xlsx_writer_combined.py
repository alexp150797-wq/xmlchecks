from pathlib import Path
from openpyxl import load_workbook
from xmlchecks.pkg.xlsx_writer_combined import write_combined_xlsx


def test_combined_writer_recommendation(tmp_path):
    rows_xml = [{
        "Имя файла": "a.ifc",
        "Файл из XML": "a.ifc",
        "CRC-32 XML": "AAAA",
        "CRC-32 IFC": "AAAA",
        "Имя совпадает": "Да",
        "CRC совпадает": "Да",
        "Статус": "OK",
        "Подробности": None,
        "recommendation": "rec",
    }]
    out = tmp_path / 'out.xlsx'
    write_combined_xlsx(rows_xml, None, None, out)
    wb = load_workbook(out)
    ws = wb['XML - IFC']
    assert ws['I2'].value == 'rec'
    sm = wb['Итого XML']
    assert sm.cell(row=2, column=2).value == 1
